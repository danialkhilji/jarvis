from pathlib import Path
import config

SKILL_NAME = "edit_excel"
SKILL_DESCRIPTION = (
    "Read or edit Excel files (.xlsx). "
    "Actions: read, write_cell, append_row, list_sheets. "
    "Returns data as a table or confirmation of changes."
)


def run(
    file_path: str,
    action: str,
    output_path: str = "",
    sheet_name: str = "",
    cell: str = "",
    value: str = "",
    row_data: str = "",
) -> str:
    """
    Work with Excel files.

    Args:
        file_path: Absolute path to the .xlsx file.
        action: One of 'read', 'write_cell', 'append_row', 'list_sheets'.
        output_path: For write actions; defaults to overwriting original.
        sheet_name: Target sheet name. Uses active sheet if empty.
        cell: Cell reference like 'A1' (for 'write_cell').
        value: Value to write (for 'write_cell').
        row_data: Comma-separated values for a new row (for 'append_row').
    """
    max_rows = 10000
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl is not installed. Run: pip install openpyxl"

    src = Path(file_path)
    if not src.exists():
        return f"Error: File not found: {file_path}"
    if src.suffix.lower() != ".xlsx":
        return "Error: Only .xlsx files are supported."

    allowed = any(str(src).startswith(d) for d in config.ALLOWED_DIRECTORIES)
    if not allowed:
        return "Error: Access denied. Path must be under an allowed directory."

    wb = openpyxl.load_workbook(str(src))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    if action == "list_sheets":
        return "Sheets: " + ", ".join(wb.sheetnames)

    if action == "read":
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(f"... (truncated at {max_rows} rows)")
                break
            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
        return "\n".join(rows) if rows else "Sheet is empty."

    dest = Path(output_path) if output_path else src

    if action == "write_cell":
        if not cell or not value:
            return "Error: 'write_cell' requires cell and value."
        ws[cell.upper()] = _coerce(value)

    elif action == "append_row":
        if not row_data:
            return "Error: 'append_row' requires row_data."
        values = [_coerce(v.strip()) for v in row_data.split(",")]
        ws.append(values)

    else:
        return f"Error: Unknown action '{action}'. Choose from: read, write_cell, append_row, list_sheets"

    wb.save(str(dest))
    return f"Excel file saved: {dest}"


def _coerce(val: str):
    try:
        return int(val)
    except ValueError:
        pass
    try:
        return float(val)
    except ValueError:
        return val
